import datetime
from io import BytesIO
from typing import List, Dict, Any, Optional
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def format_consolidated_value(val: Optional[float], item: Dict[str, Any]) -> str:
    if val is None:
        return "—"
    is_percent = item.get("tipo_exibicao") == "PERCENTAGE"
    base = item.get("percent_base", "DECIMAL")
    casas = item.get("casas_decimais")
    if casas is None:
        casas = 2 if is_percent else 4
    
    display_val = float(val)
    if is_percent and base == "DECIMAL":
        display_val = display_val * 100
        
    # Format with dot as thousands separator and comma as decimal separator
    val_str = f"{display_val:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{val_str}%" if is_percent else val_str

def generate_harvest_plan_pdf(consolidation_data: Dict[str, Any], year_harvest: str) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()

    primary_color = colors.HexColor('#0d9488')  # Teal-600
    text_color = colors.HexColor('#1e293b')     # Slate-800
    bg_light = colors.HexColor('#f8fafc')       # Slate-50
    bg_divider = colors.HexColor('#f1f5f9')     # Slate-100

    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=16, textColor=primary_color, spaceAfter=4)
    meta_style = ParagraphStyle('DocMeta', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#475569'), spaceAfter=12)
    header_cell_style = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontSize=7, fontName='Helvetica-Bold', textColor=colors.white)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=6, leading=8, textColor=text_color)
    cell_mono_style = ParagraphStyle('CellMono', parent=styles['Normal'], fontSize=6, fontName='Courier', leading=8, textColor=text_color)
    divider_style = ParagraphStyle('Divider', parent=styles['Normal'], fontSize=7, fontName='Helvetica-Bold', textColor=primary_color)

    story.append(Paragraph("UISA - Plano de Safra Consolidado", title_style))
    meta_text = f"<b>Ano Safra:</b> {year_harvest} &nbsp;&nbsp;|&nbsp;&nbsp; <b>Gerado em:</b> {datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')}"
    story.append(Paragraph(meta_text, meta_style))

    months = consolidation_data.get("months", [])
    data = consolidation_data.get("data", [])

    # Table headers
    headers = [
        Paragraph("ID", header_cell_style),
        Paragraph("Descrição", header_cell_style),
        Paragraph("Setor", header_cell_style),
        Paragraph("Un.", header_cell_style),
        Paragraph("Regra", header_cell_style)
    ]
    for m in months:
        headers.append(Paragraph(m, header_cell_style))
    headers.append(Paragraph("Acumulado", header_cell_style))

    table_data = [headers]
    t_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]

    for idx, item in enumerate(data):
        row_idx = idx + 1
        if item.get("tipo_item") == "divider":
            row = [Paragraph(item.get("label", ""), divider_style)] + [""] * (len(months) + 5)
            table_data.append(row)
            t_styles.append(('SPAN', (0, row_idx), (-1, row_idx)))
            t_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_divider))
        else:
            op_label = item.get("harvest_plan_op", "SUM") or "SUM"
            row = [
                Paragraph(item.get("variable_id", ""), cell_mono_style),
                Paragraph(item.get("nome", ""), cell_style),
                Paragraph(item.get("setor_id", ""), cell_style),
                Paragraph(item.get("unidade", ""), cell_style),
                Paragraph(op_label, cell_style)
            ]
            for m in months:
                val = item.get("monthly_values", {}).get(m)
                row.append(Paragraph(format_consolidated_value(val, item), cell_mono_style))
            
            accum = item.get("accumulated", {})
            accum_val = format_consolidated_value(accum.get("value"), item) if accum.get("status") == "OK" else accum.get("status", "PENDING")
            row.append(Paragraph(accum_val, cell_mono_style))
            table_data.append(row)
            
            if row_idx % 2 == 0:
                t_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_light))

    # Calculate column widths dynamically to fit page width (approx 720 points total)
    base_widths = [50, 110, 45, 25, 35]
    month_width = 405 / len(months) if months else 405
    col_widths = base_widths + [month_width] * len(months) + [50]

    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle(t_styles))
    story.append(t)

    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_harvest_plan_xlsx(consolidation_data: Dict[str, Any], year_harvest: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plano de Safra"

    ws.append(["UISA - PLANO DE SAFRA CONSOLIDADO"])
    ws.append([f"Ano Safra: {year_harvest}"])
    ws.append([f"Exportado em: {datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S UTC')}"])
    ws.append([])  # Spacer

    months = consolidation_data.get("months", [])
    data = consolidation_data.get("data", [])

    headers = ["ID", "Descrição", "Setor", "Unidade", "Regra Agregação"] + months + ["Acumulado"]
    ws.append(headers)

    # Styles
    title_font = Font(name="Calibri", size=11, bold=True, color="0D9488")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    divider_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    divider_font = Font(name="Calibri", size=10, bold=True, color="0D9488")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )

    # Style metadata headers
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="0D9488")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True)
    ws["A3"].font = Font(name="Calibri", size=10, italic=True)

    # Style table headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")
        cell.border = border_thin

    for item in data:
        row_idx = ws.max_row + 1
        if item.get("tipo_item") == "divider":
            ws.append([item.get("label", "")] + [""] * (len(headers) - 1))
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            
            # Style divider
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = divider_fill
                cell.font = divider_font
                cell.border = border_thin
        else:
            op_label = item.get("harvest_plan_op", "SUM") or "SUM"
            row = [
                item.get("variable_id", ""),
                item.get("nome", ""),
                item.get("setor_id", ""),
                item.get("unidade", ""),
                op_label
            ]
            for m in months:
                val = item.get("monthly_values", {}).get(m)
                row.append(format_consolidated_value(val, item))
            
            accum = item.get("accumulated", {})
            accum_val = format_consolidated_value(accum.get("value"), item) if accum.get("status") == "OK" else accum.get("status", "PENDING")
            row.append(accum_val)
            ws.append(row)

            # Style data row
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="right" if col_idx > 5 else "left", vertical="center")
                if is_even:
                    cell.fill = zebra_fill

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row < 5:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
